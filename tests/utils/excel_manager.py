"""
Excel操作工具类
用于读取和写入测试结果到Excel表格
"""

import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from typing import Dict, List, Any, Optional
from datetime import datetime


class ExcelManager:
    """Excel文件管理器"""
    
    HEADERS = [
        "id",
        "requirement",
        "package_name",
        "class_name",
        "method_name",
        "parameters",
        "return_type",
        "is_interface",
        "class_type",
        "fields",
        "dependencies",
        "structured_result",
        "parse_time_ms",
        "session_id",
        "test_code",
        "empty_method",
        "generate_time_ms",
        "compile_success",
        "compile_error",
        "fix_code_1",
        "fix_success_1",
        "fix_time_1",
        "fix_code_2",
        "fix_success_2",
        "fix_time_2",
        "fix_code_3",
        "fix_success_3",
        "fix_time_3",
        "final_success",
        "total_fix_count"
    ]
    
    def __init__(self, file_path: str):
        self.file_path = file_path
        self.workbook = None
        self.sheet = None
        
    def create(self):
        """创建新的Excel文件"""
        self.workbook = Workbook()
        self.sheet = self.workbook.active
        self.sheet.title = "Test Results"
        
        for col, header in enumerate(self.HEADERS, 1):
            cell = self.sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        self.workbook.save(self.file_path)
        
    def load(self):
        """加载现有Excel文件"""
        if not os.path.exists(self.file_path):
            self.create()
        else:
            self.workbook = load_workbook(self.file_path)
            self.sheet = self.workbook.active
            
    def get_row_count(self) -> int:
        """获取数据行数"""
        return self.sheet.max_row - 1
        
    def get_row(self, row_index: int) -> Dict[str, Any]:
        """获取指定行的数据"""
        row_data = {}
        for col, header in enumerate(self.HEADERS, 1):
            cell_value = self.sheet.cell(row=row_index + 1, column=col).value
            row_data[header] = cell_value
        return row_data
        
    def add_row(self, data: Dict[str, Any]):
        """添加新行"""
        row_index = self.sheet.max_row + 1
        for col, header in enumerate(self.HEADERS, 1):
            value = data.get(header, "")
            self.sheet.cell(row=row_index, column=col, value=value)
            
    def update_cell(self, row_index: int, column_name: str, value: Any):
        """更新指定单元格"""
        if column_name in self.HEADERS:
            col = self.HEADERS.index(column_name) + 1
            self.sheet.cell(row=row_index + 1, column=col, value=value)
            
    def update_row(self, row_index: int, data: Dict[str, Any]):
        """更新整行数据"""
        for col, header in enumerate(self.HEADERS, 1):
            if header in data:
                self.sheet.cell(row=row_index + 1, column=col, value=data[header])
                
    def save(self):
        """保存Excel文件"""
        self.workbook.save(self.file_path)
        
    def get_all_rows(self) -> List[Dict[str, Any]]:
        """获取所有行数据"""
        rows = []
        for i in range(1, self.sheet.max_row):
            rows.append(self.get_row(i))
        return rows
        
    def get_rows_by_condition(self, column_name: str, value: Any) -> List[Dict[str, Any]]:
        """根据条件获取行"""
        rows = []
        for i in range(1, self.sheet.max_row):
            row_data = self.get_row(i)
            if row_data.get(column_name) == value:
                row_data["_row_index"] = i
                rows.append(row_data)
        return rows
        
    def set_column_width(self):
        """设置列宽"""
        column_widths = {
            "id": 8,
            "requirement": 50,
            "package_name": 25,
            "class_name": 20,
            "method_name": 20,
            "parameters": 30,
            "return_type": 15,
            "is_interface": 12,
            "class_type": 15,
            "fields": 30,
            "dependencies": 30,
            "structured_result": 40,
            "parse_time_ms": 15,
            "session_id": 40,
            "test_code": 60,
            "empty_method": 40,
            "generate_time_ms": 18,
            "compile_success": 15,
            "compile_error": 40,
            "fix_code_1": 60,
            "fix_success_1": 15,
            "fix_time_1": 15,
            "fix_code_2": 60,
            "fix_success_2": 15,
            "fix_time_2": 15,
            "fix_code_3": 60,
            "fix_success_3": 15,
            "fix_time_3": 15,
            "final_success": 15,
            "total_fix_count": 18
        }
        
        for col, header in enumerate(self.HEADERS, 1):
            width = column_widths.get(header, 15)
            self.sheet.column_dimensions[self.sheet.cell(row=1, column=col).column_letter].width = width
